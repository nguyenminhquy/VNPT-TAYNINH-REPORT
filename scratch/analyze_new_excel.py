import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
wb = openpyxl.load_workbook('data sample/DATA SAMPLE THANG/1. BÁO CÁO MBB_HUNG.xlsx', data_only=True)
for sheet_name in wb.sheetnames:
    print(f"--- Sheet: {sheet_name} ---")
    sheet = wb[sheet_name]
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=3, values_only=True)):
        print(f"Row {i+1}: {row}")
