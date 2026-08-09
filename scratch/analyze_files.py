import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

for f in os.listdir('data sample/DATA SAMPLE THANG'):
    if 'MBB' in f or 'FBB' in f or 'MYTV' in f.upper() or 'ISPEED' in f.upper() or '5S' in f.upper():
        print(f'\n=== FILE: {f} ===')
        try:
            wb = openpyxl.load_workbook(f'data sample/DATA SAMPLE THANG/{f}', data_only=True)
            for sheet_name in wb.sheetnames:
                print(f'--- Sheet: {sheet_name} ---')
                sheet = wb[sheet_name]
                for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=8, min_col=1, max_col=3, values_only=True)):
                    print(f'Row {i+1}: {row}')
        except Exception as e:
            print('Error:', e)
