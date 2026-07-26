"""
Script Python xuất báo cáo Word VNPT Tây Ninh từ template.docx và các file Excel.
Thay thế hoàn toàn logic TypeScript bị lỗi.
Đọc các file Excel từ Vercel Blob URL hoặc từ thư mục local nếu chạy test.
"""

import os, sys, re, zipfile, json, shutil
import urllib.request, urllib.error
from xml.etree import ElementTree as ET
from copy import deepcopy
from datetime import datetime, date, timedelta

# ── CONSTANTS ─────────────────────────────────────────────────────────────────
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'templates', 'template.docx')
NS = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
W = f'{{{NS}}}'

# ── ISO WEEK CALCULATION ──────────────────────────────────────────────────────
def get_iso_week(d=None):
    if d is None:
        d = date.today()
    year, week, _ = d.isocalendar()
    return week, year

# ── EXCEL READER ──────────────────────────────────────────────────────────────
class ExcelReader:
    """Đọc file Excel .xlsx bằng openpyxl."""
    def __init__(self, path_or_bytes):
        import openpyxl, io
        if isinstance(path_or_bytes, (str, os.PathLike)):
            self.wb = openpyxl.load_workbook(path_or_bytes, data_only=True)
        else:
            self.wb = openpyxl.load_workbook(io.BytesIO(path_or_bytes), data_only=True)

    def sheet(self, name):
        if name in self.wb.sheetnames:
            return SheetReader(self.wb[name])
        # fuzzy match
        for s in self.wb.sheetnames:
            if s.strip().lower() == name.strip().lower():
                return SheetReader(self.wb[s])
        available = ', '.join(self.wb.sheetnames)
        raise KeyError(f'Sheet "{name}" not found. Available: {available}')

    def sheet_startswith(self, prefix):
        for name in self.wb.sheetnames:
            if name.startswith(prefix):
                return SheetReader(self.wb[name]), name
        return None, None

    @property
    def sheetnames(self):
        return self.wb.sheetnames

class SheetReader:
    def __init__(self, ws):
        self.ws = ws

    def cell(self, row, col):
        """1-indexed row, col. Returns raw value."""
        return self.ws.cell(row=row, column=col).value

    def cell_str(self, row, col):
        v = self.cell(row, col)
        if v is None:
            return ''
        return str(v).strip()

    def cell_num(self, row, col):
        v = self.cell(row, col)
        if v is None:
            return 0.0
        try:
            return float(str(v).replace('%', '').replace(',', '').strip())
        except:
            return 0.0

    def matrix(self, row_start, row_end, col_start, col_end):
        """Returns list of list of string. Inclusive 1-indexed."""
        result = []
        for r in range(row_start, row_end + 1):
            row = []
            for c in range(col_start, col_end + 1):
                v = self.cell(r, c)
                row.append('' if v is None else str(v).strip())
            result.append(row)
        return result

    def matrix_until_blank(self, row_start, col_start, col_end):
        """Returns matrix rows until a row is fully blank."""
        result = []
        r = row_start
        while True:
            row = []
            has_data = False
            for c in range(col_start, col_end + 1):
                v = self.cell(r, c)
                s = '' if v is None else str(v).strip()
                row.append(s)
                if s:
                    has_data = True
            if not has_data:
                break
            result.append(row)
            r += 1
        return result

# ── FORMATTING HELPERS ────────────────────────────────────────────────────────
def fmt_pct(val, digits=2, stored_as_pct=False):
    """Format số thành phần trăm. Excel lưu 0.9524 hoặc 95.24."""
    try:
        n = float(str(val).replace('%', '').replace(',', '').strip())
        if not stored_as_pct and abs(n) <= 1.5:
            n *= 100
        return f'{n:.{digits}f}%'
    except:
        return str(val) if val else ''

def fmt_int(val, dash_zero=False):
    try:
        n = round(float(str(val).replace(',', '').strip()))
        if dash_zero and n == 0:
            return '-'
        return f'{n:,}'
    except:
        return str(val) if val else ''

def fmt_dec(val, digits=2):
    try:
        n = float(str(val).replace(',', '').strip())
        return f'{n:.{digits}f}'
    except:
        return str(val) if val else ''

def clean(val):
    if val is None:
        return ''
    return str(val).strip()

def evaluate_target(val, target=99.0):
    try:
        n = float(str(val).replace('%', '').replace(',', '').strip())
        if abs(n) <= 1.5:
            n *= 100
        return 'Đạt' if n >= target else 'Không đạt'
    except:
        return ''

# ── DOCX MODIFIER ─────────────────────────────────────────────────────────────
class DocxModifier:
    """Sửa file DOCX qua XML, tương thích đầy đủ với template.docx."""

    def __init__(self, template_path):
        self.zip = zipfile.ZipFile(template_path, 'r')
        self.xml_bytes = self.zip.read('word/document.xml')
        ET.register_namespace('', NS)
        for prefix, uri in [
            ('wpc', 'http://schemas.microsoft.com/office/word/2010/wordprocessingCanvas'),
            ('cx', 'http://schemas.microsoft.com/office/drawing/2014/chartex'),
            ('m', 'http://schemas.openxmlformats.org/officeDocument/2006/math'),
            ('mc', 'http://schemas.openxmlformats.org/markup-compatibility/2006'),
            ('o', 'urn:schemas-microsoft-com:office:office'),
            ('oel', 'http://schemas.microsoft.com/office/2019/extlst'),
            ('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'),
            ('m14', 'http://schemas.microsoft.com/office/math/2010/mathml/element'),
            ('v', 'urn:schemas-microsoft-com:vml'),
            ('w', NS),
            ('w14', 'http://schemas.microsoft.com/office/word/2010/wordml'),
            ('w15', 'http://schemas.microsoft.com/office/word/2012/wordml'),
            ('w16', 'http://schemas.microsoft.com/office/word/2018/wordml/cid'),
            ('w16cex', 'http://schemas.microsoft.com/office/word/2018/wordml/cex'),
            ('w16sdtdh', 'http://schemas.microsoft.com/office/word/2020/wordml/sdtdatahash'),
            ('wne', 'http://schemas.microsoft.com/office/word/2006/wordml'),
            ('wp', 'http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing'),
            ('wps', 'http://schemas.microsoft.com/office/word/2010/wordprocessingShape'),
        ]:
            try:
                ET.register_namespace(prefix, uri)
            except:
                pass

        self.tree = ET.parse(zipfile.ZipFile(template_path, 'r').open('word/document.xml'))
        self.root = self.tree.getroot()
        self.body = self.root.find(f'{W}body')

    def _get_text(self, elem):
        return ''.join(t.text or '' for t in elem.iter(f'{W}t')).strip()

    def get_paragraphs(self):
        """Trả về list (index, elem) của các <w:p> trực tiếp trong body."""
        result = []
        for i, child in enumerate(self.body):
            if child.tag == f'{W}p':
                result.append(child)
        return result

    def get_tables(self):
        """Trả về list tất cả <w:tbl> trong body."""
        return [child for child in self.body if child.tag == f'{W}tbl']

    def _set_run_text(self, run, text):
        """Xóa các <w:t> cũ và set text mới, giữ properties."""
        rpr = run.find(f'{W}rPr')
        for child in list(run):
            if child.tag != f'{W}rPr':
                run.remove(child)
        t_elem = ET.SubElement(run, f'{W}t')
        t_elem.text = str(text)
        if text and (text[0] == ' ' or text[-1] == ' '):
            t_elem.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')

    def replace_para_text(self, para, text):
        """Thay nội dung text của một paragraph, giữ formatting của run đầu tiên."""
        ppr = para.find(f'{W}pPr')

        # Lấy rPr từ run đầu tiên
        rpr = None
        first_run = para.find(f'{W}r')
        if first_run is not None:
            rpr = first_run.find(f'{W}rPr')

        # Xóa tất cả children ngoài pPr
        for child in list(para):
            if child.tag != f'{W}pPr':
                para.remove(child)

        # Thêm run mới
        run = ET.SubElement(para, f'{W}r')
        if rpr is not None:
            run.append(deepcopy(rpr))
        t_elem = ET.SubElement(run, f'{W}t')
        t_elem.text = str(text)
        if text and (str(text)[0] == ' ' or str(text)[-1] == ' '):
            t_elem.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')

    def replace_para_by_pattern(self, pattern, replacement):
        """Thay paragraph đầu tiên khớp pattern."""
        paragraphs = list(self.body.iter(f'{W}p'))
        for para in paragraphs:
            text = self._get_text(para)
            if isinstance(pattern, str):
                match = pattern in text
                new_text = text.replace(pattern, replacement) if match else text
            else:
                match = bool(pattern.search(text))
                new_text = pattern.sub(replacement, text) if match else text
            if match:
                self.replace_para_text(para, new_text)
                return True
        return False

    def replace_all_paras_by_pattern(self, pattern, replacement):
        """Thay tất cả paragraphs khớp pattern."""
        paragraphs = list(self.body.iter(f'{W}p'))
        replaced = 0
        for para in paragraphs:
            text = self._get_text(para)
            if isinstance(pattern, str):
                match = pattern in text
                new_text = text.replace(pattern, replacement) if match else text
            else:
                match = bool(pattern.search(text))
                new_text = pattern.sub(replacement, text) if match else text
            if match:
                self.replace_para_text(para, new_text)
                replaced += 1
        return replaced

    def replace_para_by_index(self, index, text):
        """Thay paragraph theo index trong body-level paragraphs."""
        paras = self.get_paragraphs()
        if index < len(paras):
            self.replace_para_text(paras[index], text)

    def _get_table_rows(self, tbl):
        return [child for child in tbl if child.tag == f'{W}tr']

    def _get_row_cells(self, row):
        return [child for child in row if child.tag == f'{W}tc']

    def _set_cell_text(self, cell, text):
        """Thay text trong cell, giữ formatting."""
        paras = [child for child in cell if child.tag == f'{W}p']
        if not paras:
            para = ET.SubElement(cell, f'{W}p')
            paras = [para]

        # Dùng para đầu tiên
        para = paras[0]
        self.replace_para_text(para, str(text))

        # Xóa các para thừa
        for extra_para in paras[1:]:
            cell.remove(extra_para)

    def write_table_matrix(self, tbl, matrix, start_row=0):
        """Ghi matrix (list of list) vào bảng từ start_row."""
        rows = self._get_table_rows(tbl)
        for ri, row_data in enumerate(matrix):
            target_ri = start_row + ri
            if target_ri >= len(rows):
                break
            row = rows[target_ri]
            cells = self._get_row_cells(row)
            for ci, val in enumerate(row_data):
                if ci >= len(cells):
                    break
                cell = cells[ci]
                # Skip vertically merged continuation cells
                tcPr = cell.find(f'{W}tcPr')
                if tcPr is not None:
                    vMerge = tcPr.find(f'{W}vMerge')
                    if vMerge is not None:
                        w_val = vMerge.get(f'{W}val')
                        if not w_val or w_val == 'continue':
                            continue
                self._set_cell_text(cell, val)

    def resize_table_rows(self, tbl, desired_rows):
        """Thêm hoặc xóa hàng để đạt desired_rows."""
        rows = self._get_table_rows(tbl)
        current = len(rows)
        while current < desired_rows:
            last_row = rows[-1]
            new_row = deepcopy(last_row)
            # Xóa text trong cells mới
            for cell in self._get_row_cells(new_row):
                self._set_cell_text(cell, '')
            tbl.append(new_row)
            rows.append(new_row)
            current += 1
        while current > desired_rows:
            tbl.remove(rows[-1])
            rows.pop()
            current -= 1

    def find_table_by_internal_text(self, keyword):
        """Tìm bảng chứa keyword bên trong (tìm từ cuối về đầu)."""
        tables = self.get_tables()
        for tbl in reversed(tables):
            tbl_text = ''.join(t.text or '' for t in tbl.iter(f'{W}t'))
            if keyword.lower() in tbl_text.lower():
                return tbl
        return None

    def find_table_by_preceding_text(self, keyword):
        """Tìm bảng đứng ngay sau paragraph chứa keyword."""
        children = list(self.body)
        found = False
        for child in children:
            if child.tag == f'{W}p':
                text = self._get_text(child)
                if keyword.lower() in text.lower():
                    found = True
            elif found and child.tag == f'{W}tbl':
                return child
        return None

    def save(self, output_path):
        """Lưu file docx ra output_path."""
        import io, shutil
        # Serialize XML
        # Đăng ký namespace để tránh ns0:, ns1:...
        xml_str = ET.tostring(self.root, encoding='unicode', xml_declaration=False)
        xml_bytes = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + xml_str).encode('utf-8')

        # Copy template zip và thay word/document.xml
        with zipfile.ZipFile(TEMPLATE_PATH, 'r') as zin:
            with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    if item.filename == 'word/document.xml':
                        zout.writestr(item, xml_bytes)
                    else:
                        zout.writestr(item, zin.read(item.filename))

    def save_bytes(self):
        """Trả về nội dung file docx dưới dạng bytes."""
        import io
        xml_str = ET.tostring(self.root, encoding='unicode', xml_declaration=False)
        xml_bytes = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + xml_str).encode('utf-8')
        buf = io.BytesIO()
        with zipfile.ZipFile(TEMPLATE_PATH, 'r') as zin:
            with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    if item.filename == 'word/document.xml':
                        zout.writestr(item, xml_bytes)
                    else:
                        zout.writestr(item, zin.read(item.filename))
        return buf.getvalue()


# ── EXTRACT FUNCTIONS ─────────────────────────────────────────────────────────

def update_mbb_fbb_mytv(doc: DocxModifier, mbb: ExcelReader, fbb: ExcelReader, mytv: ExcelReader):
    """Cập nhật dữ liệu MBB, FBB, MyTV vào template."""

    # ── Bảng 1: STT | Đơn vị | QoS | QoE (sheet Kết quả chung, rows 4-6, cols A-D)
    kq_sheet = mbb.sheet('Kết quả chung')
    kq_matrix = kq_sheet.matrix_until_blank(5, 1, 4)   # Row 5+ (bỏ header), cột A-D
    # Format cột QoS/QoE
    for row in kq_matrix:
        row[2] = fmt_dec(row[2]) if row[2] else ''
        row[3] = fmt_dec(row[3]) if row[3] else ''
    tbl1 = doc.find_table_by_internal_text('STTĐơn vịQoSQoE')
    if tbl1 is None:
        tbl1 = doc.find_table_by_internal_text('STT')
    if tbl1 and kq_matrix:
        doc.resize_table_rows(tbl1, len(kq_matrix))
        doc.write_table_matrix(tbl1, kq_matrix)

    # ── Bảng 2: Tỉnh | QoS MBB | QoE MBB (So sánh các tỉnh, rows 4+)
    ss_sheet = mbb.sheet('So sánh các tỉnh')
    ss_matrix = ss_sheet.matrix_until_blank(4, 1, 3)
    tbl2 = doc.find_table_by_internal_text('TỉnhQoS MBBQoE MBB')
    if tbl2 and ss_matrix:
        doc.resize_table_rows(tbl2, len(ss_matrix))
        doc.write_table_matrix(tbl2, ss_matrix)

    # ── Bảng 3: Kết quả chi tiết (phức hợp MBB+FBB+MyTV)
    # MBB detail: Kết quả chi tiết rows 4+, cols B-H (2-8)
    ct_sheet = mbb.sheet('Kết quả chi tiết')
    mbb_detail = ct_sheet.matrix_until_blank(4, 2, 8)
    # Pad to 9 rows max
    while len(mbb_detail) < 9:
        mbb_detail.append([''] * 7)
    mbb_detail = mbb_detail[:9]

    # FBB detail: Thông tin chung rows 2+, cols B-H (2-8)
    try:
        fbb_common = fbb.sheet('Thông tin chung')
        fbb_detail = fbb_common.matrix_until_blank(2, 2, 8)
        while len(fbb_detail) < 16:
            fbb_detail.append([''] * 7)
        fbb_detail = fbb_detail[:16]
    except:
        fbb_detail = [[''] * 7] * 16

    # MyTV detail: Sheet1 rows 2+, cols B-H (2-8)
    try:
        mytv_sheet = mytv.sheet('Sheet1')
        mytv_raw = mytv_sheet.matrix_until_blank(2, 2, 9)  # cols B-I
        mytv_detail = []
        for row in mytv_raw:
            total = row[6] if len(row) > 6 else ''
            mytv_detail.append([
                clean(row[0]), clean(row[1]),
                clean(row[3]) if len(row) > 3 else '',
                clean(row[4]) if len(row) > 4 else '',
                clean(row[5]) if len(row) > 5 else '',
                clean(total),
                evaluate_target(total),
                clean(row[7]) if len(row) > 7 else ''
            ])
        while len(mytv_detail) < 14:
            mytv_detail.append([''] * 8)
        mytv_detail = mytv_detail[:14]
    except:
        mytv_detail = [[''] * 8] * 14

    tbl3 = doc.find_table_by_internal_text('Thành phầnĐiểm thành phầnTổng')
    if tbl3:
        doc.write_table_matrix(tbl3, mbb_detail, start_row=3)
        doc.write_table_matrix(tbl3, fbb_detail, start_row=13)
        doc.write_table_matrix(tbl3, mytv_detail, start_row=30)

    # ── Bảng 4: Giải trình QoS
    qos_explain_sheet = mbb.sheet('Giải trình QoS')
    qos_explain = qos_explain_sheet.matrix_until_blank(5, 1, 4)
    tbl4 = doc.find_table_by_preceding_text('Chỉ số QoS MBB')
    if tbl4 and qos_explain:
        doc.resize_table_rows(tbl4, len(qos_explain))
        doc.write_table_matrix(tbl4, qos_explain)

    # ── Bảng 5: Giải trình QoE
    qoe_explain_sheet = mbb.sheet('Giải trình QoE')
    qoe_explain = qoe_explain_sheet.matrix_until_blank(3, 1, 4)
    tbl5 = doc.find_table_by_preceding_text('Chỉ số QoE MBB')
    if tbl5 and qoe_explain:
        doc.resize_table_rows(tbl5, len(qoe_explain))
        doc.write_table_matrix(tbl5, qoe_explain)

    # ── Bảng 6: Dự kiến tuần kế tiếp
    plan_sheet, plan_sheet_name = mbb.sheet_startswith('Dự kiến tuần')
    plan_data = plan_sheet.matrix_until_blank(4, 1, 4) if plan_sheet else []
    tbl6 = doc.find_table_by_preceding_text('Công việc dự kiến tuần')
    if tbl6 and plan_data:
        doc.resize_table_rows(tbl6, len(plan_data))
        doc.write_table_matrix(tbl6, plan_data)

    # Update "Công việc dự kiến tuần X" label
    if plan_sheet_name:
        m = re.search(r'(\d+)$', plan_sheet_name)
        if m:
            doc.replace_para_by_pattern(
                re.compile(r'Công việc dự kiến tuần \d+:'),
                f'Công việc dự kiến tuần {m.group(1)}:'
            )

    # ── Bảng 7: Phản ánh khách hàng
    pakh_sheet = mbb.sheet('Phản ánh khách hàng (PAKH)')
    pakh_data = pakh_sheet.matrix_until_blank(4, 2, 3)
    tbl7 = doc.find_table_by_preceding_text('Kết quả thực hiện:')
    if tbl7 and pakh_data:
        doc.resize_table_rows(tbl7, len(pakh_data))
        doc.write_table_matrix(tbl7, pakh_data)

    # Cập nhật "Thời gian lấy báo cáo" từ PAKH
    cutoff_text = ''
    for row in pakh_data[1:]:
        if row and 'đến' in str(row[1]).lower():
            cutoff_text = row[1]
            break
    m = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', cutoff_text)
    if m:
        parts = m.group(1).split('/')
        d, mo, y = parts[0].zfill(2), parts[1].zfill(2), parts[2]
        doc.replace_para_by_pattern(
            re.compile(r'Thời gian lấy báo cáo: từ \d{2}/\d{2}/\d{4} – \d{2}/\d{2}/\d{4}'),
            f'Thời gian lấy báo cáo: từ 01/{mo}/{y} – {d}/{mo}/{y}'
        )

    # ── FBB Bảng 8: Nguyên nhân/Giải pháp
    fbb_qos_sheet = fbb.sheet('Chi tiết QoS FBB')
    tbl8 = doc.find_table_by_internal_text('Nguyên nhânGiải pháp')
    if tbl8:
        data8 = fbb_qos_sheet.matrix(2, 2, 2, 2)
        doc.resize_table_rows(tbl8, len(data8))
        doc.write_table_matrix(tbl8, data8)

    # ── FBB Bảng 9: Nghẽn BRCĐ
    tbl9 = doc.find_table_by_internal_text('NgàyOLTSố lượng Uplink')
    if tbl9:
        data9 = fbb_qos_sheet.matrix_until_blank(6, 1, 7)
        if data9:
            doc.resize_table_rows(tbl9, len(data9))
            doc.write_table_matrix(tbl9, data9)

    # ── FBB Bảng 10: Chỉ số theo THT
    tbl10 = doc.find_table_by_internal_text('STTTHTFBB QoSĐạt/Chưa đạt')
    if tbl10:
        data10 = fbb_qos_sheet.matrix_until_blank(10, 1, 4)
        if data10:
            doc.resize_table_rows(tbl10, len(data10))
            doc.write_table_matrix(tbl10, data10)

    # ── FBB Bảng 11: Chỉ số theo TTVT
    tbl11 = doc.find_table_by_internal_text('STTTHTTTVTFBB QoSĐạt/Chưa đạt')
    if tbl11:
        data11 = fbb_qos_sheet.matrix_until_blank(21, 1, 5)
        if data11:
            doc.resize_table_rows(tbl11, len(data11))
            doc.write_table_matrix(tbl11, data11)

    # ── FBB Bảng 12: Suy hao thuê bao
    tbl12 = doc.find_table_by_internal_text('Thuê bao suy haoTỉ lệ suy hao')
    if tbl12:
        suy_hao = fbb.sheet('Suy hao thuê bao')
        data12 = suy_hao.matrix_until_blank(3, 1, 7)
        if data12:
            doc.resize_table_rows(tbl12, len(data12))
            doc.write_table_matrix(tbl12, data12)


def update_mll(doc: DocxModifier, mll: ExcelReader) -> str:
    """Cập nhật dữ liệu MLL."""
    sheet = mll.sheet('BC MLL tuần')
    raw = sheet.matrix(2, 11, 1, 18)  # rows 2-11, cols 1-18

    title = raw[0][0] if raw and raw[0] else ''
    week_match = re.search(r'TUẦN\s+(\d+)', title.upper())
    week = week_match.group(1) if week_match else ''

    # Bảng MLL: rows 3-11 (index 2-10 trong raw, skip row1=title, row2=header)
    SOURCE_COLS = [0, 1, 2, 3, 4, 5, 7, 8, 9, 11, 12, 13, 15, 16, 17]
    matrix = [[title]]  # title row
    for row in raw[3:11]:  # rows 5-11 in excel (0-indexed 3-10 in raw)
        target = []
        for ti, si in enumerate(SOURCE_COLS):
            v = row[si] if si < len(row) else ''
            if ti in (0, 1):
                target.append(clean(v))
            elif ti == 2:
                target.append(fmt_int(v))
            elif 3 <= ti <= 11:
                target.append(fmt_int(v, dash_zero=True))
            elif ti in (12, 13):
                target.append(fmt_int(v))
            else:
                target.append(fmt_dec(v))
        matrix.append(target)

    tbl_mll = doc.find_table_by_internal_text('THỜI GIAN MẤT LIÊN LẠC MẠNG DI ĐỘNG')
    if tbl_mll:
        # Set title cell
        rows = doc._get_table_rows(tbl_mll)
        if rows:
            first_cells = doc._get_row_cells(rows[0])
            if first_cells:
                doc._set_cell_text(first_cells[0], title)
        doc.write_table_matrix(tbl_mll, matrix[1:], start_row=3)

    # overall = row 4 in excel (0-indexed 2 in raw)
    overall = raw[2] if len(raw) > 2 else []
    teams_raw = raw[3:10]  # rows 5-11

    def asn(v):
        try:
            return float(str(v).replace(',', '').strip()) if v else 0.0
        except:
            return 0.0

    total_mll = asn(overall[15]) if len(overall) > 15 else 0.0
    avg_mll = asn(overall[17]) if len(overall) > 17 else 0.0
    teams = [{'name': clean(r[1]) if len(r) > 1 else '', 'avg': asn(r[17]) if len(r) > 17 else 0.0} for r in teams_raw]

    cause_power = sum(asn(overall[i]) for i in [3, 7, 11] if i < len(overall))
    cause_equip = sum(asn(overall[i]) for i in [4, 8, 12] if i < len(overall))
    cause_trans = sum(asn(overall[i]) for i in [5, 9, 13] if i < len(overall))

    doc.replace_para_by_pattern(
        re.compile(r'Tổng thời gian mất liên lạc:.*'),
        f'Tổng thời gian mất liên lạc: {fmt_int(total_mll)} phút.'
    )
    doc.replace_para_by_pattern(
        re.compile(r'MLL trung bình/1 BTS:.*'),
        f'MLL trung bình/1 BTS: {avg_mll:.2f} phút.'
    )

    if week:
        doc.replace_para_by_pattern(
            re.compile(r'Đánh giá thời gian mất liên lạc vô tuyến tuần \d+:'),
            f'Đánh giá thời gian mất liên lạc vô tuyến tuần {week}:'
        )
        doc.replace_para_by_pattern(
            re.compile(r'Nguyên nhân chi tiết các trạm MLL trong tuần \d+ năm \d+.*'),
            f'Nguyên nhân chi tiết các trạm MLL trong tuần {week} năm 2026 và các đánh giá, giải pháp khắc phục (Theo phụ lục 01 đính kèm)'
        )
        doc.replace_para_by_pattern(
            re.compile(r'GIẢI TRÌNH NGUYÊN NHÂN MẤT LIÊN LẠC TRẠM TUẦN \d+'),
            f'GIẢI TRÌNH NGUYÊN NHÂN MẤT LIÊN LẠC TRẠM TUẦN {week}'
        )

    achieved = sum(1 for t in teams if t['avg'] <= 3.0)
    doc.replace_para_by_pattern(
        re.compile(r'\d+/7 THT có thời gian mất liên lạc đáp ứng chỉ tiêu.*'),
        f'{achieved}/7 THT có thời gian mất liên lạc đáp ứng chỉ tiêu của VTT (≤3 phút).'
    )

    highest = sorted(teams, key=lambda t: t['avg'], reverse=True)[:3]
    highest_text = ', '.join(f"THT {t['name']} ({t['avg']:.2f} phút/1 trạm)" for t in highest)
    doc.replace_para_by_pattern(
        re.compile(r'Thời gian mất liên lạc trung bình trên 1 trạm BTS cao nhất:.*'),
        f'Thời gian mất liên lạc trung bình trên 1 trạm BTS cao nhất: {highest_text}.'
    )

    total_safe = total_mll if total_mll > 0 else 1.0
    doc.replace_para_by_pattern(
        re.compile(r'MLL do lỗi nguồn \(\d+%\)'),
        f'MLL do lỗi nguồn ({round(cause_power / total_safe * 100)}%)'
    )
    doc.replace_para_by_pattern(
        re.compile(r'MLL do lỗi thiết bị \(\d+%\)'),
        f'MLL do lỗi thiết bị ({round(cause_equip / total_safe * 100)}%)'
    )
    doc.replace_para_by_pattern(
        re.compile(r'MLL do lỗi truyền dẫn \(\d+%\)'),
        f'MLL do lỗi truyền dẫn ({round(cause_trans / total_safe * 100)}%)'
    )
    return week


def update_ispeed(doc: DocxModifier, ispeed: ExcelReader):
    """Cập nhật dữ liệu i-Speed."""
    sheet = ispeed.sheet('Báo cáo')
    raw = sheet.matrix(2, 9, 1, 11)  # rows 2-9, cols 1-11

    matrix = []
    for ri, row in enumerate(raw):
        if ri == 0:
            matrix.append([clean(v) for v in row])
            continue
        matrix.append([
            clean(row[0]), clean(row[1]),
            fmt_int(row[2]), fmt_int(row[3]), fmt_int(row[4]),
            fmt_pct(row[5], stored_as_pct=True),
            fmt_int(row[6]), fmt_int(row[7]),
            fmt_pct(row[8], stored_as_pct=True),
            fmt_int(row[9]),
            fmt_pct(row[10], stored_as_pct=True)
        ])

    tbl = doc.find_table_by_internal_text('Tỉ lệ hoàn thành i-Speed')
    if tbl and matrix:
        doc.resize_table_rows(tbl, len(matrix))
        doc.write_table_matrix(tbl, matrix)

    # Cập nhật ngày báo cáo
    date_val = sheet.cell_str(12, 2)
    if date_val:
        doc.replace_para_by_pattern(
            re.compile(r'Thời gian lấy báo cáo: \d{2}/\d{2}/\d{4}$'),
            f'Thời gian lấy báo cáo: {date_val}'
        )

    # Cập nhật summary text
    total_row = raw[-1] if raw else []
    if total_row:
        ispeed_done = total_row[4] if len(total_row) > 4 else 0
        ispeed_goal = total_row[3] if len(total_row) > 3 else 0
        ispeed_pct  = total_row[5] if len(total_row) > 5 else 0
        st_done     = total_row[7] if len(total_row) > 7 else 0
        st_goal     = total_row[6] if len(total_row) > 6 else 0
        st_pct      = total_row[8] if len(total_row) > 8 else 0
        g5_done     = total_row[9] if len(total_row) > 9 else 0
        g5_total    = total_row[7] if len(total_row) > 7 else 0
        g5_pct      = total_row[10] if len(total_row) > 10 else 0

        doc.replace_para_by_pattern(
            re.compile(r'Công tác đo kiểm i-Speed đã thực hiện.*'),
            f'Công tác đo kiểm i-Speed đã thực hiện {fmt_int(ispeed_done)}/{fmt_int(ispeed_goal)} mẫu, đạt {fmt_pct(ispeed_pct, stored_as_pct=True)}/Tháng kế hoạch.'
        )
        doc.replace_para_by_pattern(
            re.compile(r'Công tác đo kiểm SpeedTest đã thực hiện.*'),
            f'Công tác đo kiểm SpeedTest đã thực hiện {fmt_int(st_done)}/{fmt_int(st_goal)} mẫu, đạt {fmt_pct(st_pct, stored_as_pct=True)}/Tháng kế hoạch.'
        )
        doc.replace_para_by_pattern(
            re.compile(r'Kết quả mẫu đo 5G SpeedTest đã thực hiện.*'),
            f'Kết quả mẫu đo 5G SpeedTest đã thực hiện {fmt_int(g5_done)}/{fmt_int(g5_total)} mẫu, đạt {fmt_pct(g5_pct, stored_as_pct=True)}/Tổng mẫu đã đo.'
        )


def update_5s(doc: DocxModifier, five_s: ExcelReader):
    """Cập nhật dữ liệu 5S nhà trạm."""
    import openpyxl
    sheet = five_s.sheet('Sheet1')
    ws = sheet.ws

    def find_section_start(keyword):
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and keyword.upper() in str(cell.value).upper():
                    return cell.row + 1  # Data bắt đầu sau header
        return -1

    def read_section(start_row, num_rows=8):
        if start_row < 0:
            return []
        matrix = []
        for r in range(start_row, start_row + num_rows + 2):
            row = [sheet.cell(r, c) for c in range(1, 7)]
            has_data = any(v for v in row[1:])
            if not has_data:
                continue
            matrix.append([
                clean(row[0]), clean(row[1]),
                fmt_int(row[2]), fmt_int(row[3]), fmt_int(row[4]),
                fmt_pct(row[5], stored_as_pct=True)
            ])
        return matrix

    station_start = find_section_start('5S NHÀ TRẠM')
    ac_start = find_section_start('VỆ SINH MÁY LẠNH')
    ap_start = find_section_start('5S AP/OTB')
    survey_start = find_section_start('KHẢO SÁT PHỤ TRỢ')

    station_matrix = read_section(station_start)
    ac_matrix = read_section(ac_start)
    ap_matrix = read_section(ap_start)
    survey_matrix = read_section(survey_start)

    tbl_station = doc.find_table_by_preceding_text('Tiến độ 5S nhà trạm')
    if tbl_station and station_matrix:
        doc.write_table_matrix(tbl_station, station_matrix)

    tbl_ap = doc.find_table_by_preceding_text('Tiến độ 5S AP/OTB')
    if tbl_ap and ap_matrix:
        doc.write_table_matrix(tbl_ap, ap_matrix)

    tbl_ac = doc.find_table_by_preceding_text('Tiến độ vệ sinh máy lạnh')
    if tbl_ac and ac_matrix:
        doc.write_table_matrix(tbl_ac, ac_matrix)

    # Cập nhật ngày
    today = date.today()
    date_str = f'{today.day:02d}/{today.month:02d}/{today.year}'
    doc.replace_all_paras_by_pattern('Thời gian lấy báo cáo:', f'Thời gian lấy báo cáo: {date_str}')


def update_xlsc(doc: DocxModifier, xlsc: ExcelReader) -> dict:
    """Cập nhật dữ liệu XLSC."""
    mappings = [
        ('XLSC MANE',     '1. Kết quả phiếu XLSC MANE',    '2. Kết quả phiếu XLSC MANE'),
        ('XLSC ACCESS',   '2. Kết quả phiếu XLSC ACCESS',   'Kết quả phiếu XLSC ACCESS'),
        ('XLSC VÔ TUYẾN', '3. Kết quả phiếu XLSC VÔ TUYẾN', 'Kết quả phiếu XLSC vô tuyến'),
    ]
    report_month = None

    for sheet_name, prefix, para_prefix in mappings:
        try:
            sheet = xlsc.sheet(sheet_name)
        except:
            continue

        raw = sheet.matrix(2, 11, 1, 10)
        matrix = []
        for ri, row in enumerate(raw):
            if ri <= 1:
                matrix.append([clean(v) for v in row])
                continue
            matrix.append([
                clean(row[0]),
                fmt_int(row[1]), fmt_int(row[2]), fmt_int(row[3]), fmt_int(row[4]),
                fmt_pct(row[5], stored_as_pct=True),
                fmt_int(row[6]), fmt_int(row[7]), fmt_int(row[8]),
                fmt_pct(row[9], stored_as_pct=True)
            ])

        tbl = doc.find_table_by_preceding_text(prefix)
        if tbl and matrix:
            doc.resize_table_rows(tbl, len(matrix))
            doc.write_table_matrix(tbl, matrix)

        # Parse title để lấy date range
        title_val = sheet.cell_str(1, 1)
        m = re.search(r'\((\d{2})-(\d{2})-(\d{4})\s*-\s*(\d{2})-(\d{2})-(\d{4})\)', title_val)
        if m:
            start = f'{m.group(1)}/{m.group(2)}/{m.group(3)}'
            end   = f'{m.group(4)}/{m.group(5)}/{m.group(6)}'
            month = int(m.group(5))
            year  = int(m.group(6))
            report_month = {'start': start, 'end': end, 'month': month, 'year': year}

        # Total row (row 11, index 9)
        total_row = raw[9] if len(raw) > 9 else []

        # Tìm paragraph chứa prefix và cập nhật các paragraph tiếp theo
        paras = doc.get_paragraphs()
        found_idx = -1
        for i, para in enumerate(paras):
            text = doc._get_text(para)
            if para_prefix.lower() in text.lower():
                found_idx = i
                break

        if found_idx >= 0 and report_month:
            for i in range(found_idx, min(found_idx + 20, len(paras))):
                text = doc._get_text(paras[i])
                if 'Kỳ báo cáo:' in text:
                    doc.replace_para_text(paras[i], f'Kỳ báo cáo: {report_month["start"]} – {report_month["end"]}')
                elif 'Tổng phiếu giao:' in text:
                    doc.replace_para_text(paras[i], f'Tổng phiếu giao: {fmt_int(total_row[1]) if total_row else ""} phiếu')
                elif text.startswith('Hoàn thành:') and '/' in text:
                    doc.replace_para_text(paras[i], f'Hoàn thành: {fmt_int(total_row[2]) if total_row else ""}/{fmt_int(total_row[1]) if total_row else ""} phiếu')
                elif 'Hoàn thành đúng hạn:' in text:
                    doc.replace_para_text(paras[i], f'Hoàn thành đúng hạn: {fmt_int(total_row[3]) if total_row else ""} phiếu')
                elif 'Hoàn thành quá hạn:' in text:
                    doc.replace_para_text(paras[i], f'Hoàn thành quá hạn: {fmt_int(total_row[4]) if total_row else ""} phiếu')
                elif 'Tỉ lệ đúng hạn:' in text:
                    doc.replace_para_text(paras[i], f'Tỉ lệ đúng hạn: {fmt_pct(total_row[5], stored_as_pct=True) if total_row else ""}')
                elif 'Phiếu tồn quá hạn:' in text:
                    doc.replace_para_text(paras[i], f'Phiếu tồn quá hạn: {fmt_int(total_row[8]) if total_row else ""} phiếu')

    if report_month:
        doc.replace_para_by_pattern(
            re.compile(r'KẾT QUẢ THỰC HIỆN PHIẾU SỰ CỐ CHUYÊN ĐỀ 5 THÁNG \d+ NĂM \d+:'),
            f'KẾT QUẢ THỰC HIỆN PHIẾU SỰ CỐ CHUYÊN ĐỀ 5 THÁNG {report_month["month"]} NĂM {report_month["year"]}:'
        )
    return report_month or {}


def update_appendix(doc: DocxModifier, appendix: ExcelReader):
    """Cập nhật Phụ lục 1 - MLL."""
    sheet = appendix.sheet('Báo Cáo Sự Cố Trạm')
    data = sheet.matrix_until_blank(5, 1, 10)  # Rows 5+ until blank, cols 1-10

    tbl = doc.find_table_by_preceding_text('GIẢI TRÌNH NGUYÊN NHÂN MẤT LIÊN LẠC TRẠM')
    if tbl and data:
        doc.resize_table_rows(tbl, len(data))
        doc.write_table_matrix(tbl, data)


def replace_report_week(doc: DocxModifier, week_from_mll: str = ''):
    """Cập nhật số tuần và ngày tháng trong toàn bộ tài liệu."""
    today = date.today()
    current_week, current_year = get_iso_week(today)
    week_str = str(current_week)
    day_str = f'{today.day:02d}'
    month_str = f'{today.month:02d}'

    # Xác định tuần kế hoạch từ MLL hoặc tự tính
    plan_week = str(int(week_str) + 1) if week_from_mll == week_str else (week_from_mll or str(current_week + 1))

    # Cập nhật ngày trong header table (Tây Ninh, ngày X tháng Y năm Z)
    tables = doc.get_tables()
    for tbl in tables:
        tbl_text = ''.join(t.text or '' for t in tbl.iter(f'{W}t'))
        if 'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM' in tbl_text:
            for para in tbl.iter(f'{W}p'):
                text = doc._get_text(para)
                if 'Tây Ninh, ngày' in text:
                    doc.replace_para_text(para, f'Tây Ninh, ngày {day_str} tháng {month_str} năm {current_year}')

    # Cập nhật tiêu đề báo cáo và body text
    doc.replace_para_by_pattern(
        re.compile(r'V/v thực hiện công việc trọng tâm trong tuần \d+ năm \d+'),
        f'V/v thực hiện công việc trọng tâm trong tuần {week_str} năm {current_year}'
    )
    doc.replace_para_by_pattern(
        re.compile(r'và kế hoạch thực hiện nhiệm vụ tuần \d+'),
        f'và kế hoạch thực hiện nhiệm vụ tuần {plan_week}'
    )
    doc.replace_para_by_pattern(
        re.compile(r'Trung tâm Hạ tầng báo cáo kết quả thực hiện công việc trọng tâm trong tuần \d+ năm \d+ như sau:'),
        f'Trung tâm Hạ tầng báo cáo kết quả thực hiện công việc trọng tâm trong tuần {week_str} năm {current_year} như sau:'
    )
    doc.replace_para_by_pattern(
        re.compile(r'Trên đây là báo cáo kết quả thực hiện công việc tuần \d+ năm \d+\.'),
        f'Trên đây là báo cáo kết quả thực hiện công việc tuần {week_str} năm {current_year}.'
    )


# ── MAIN ENTRY POINT ──────────────────────────────────────────────────────────

def generate_report(excel_paths: dict, output_path: str):
    """
    excel_paths: dict của {key: path_or_bytes}
    keys: mbb, fbb, mytv, mll, ispeed, 5s, xlsc, appendix
    output_path: đường dẫn file .docx kết quả
    """
    print('  [1/9] Đọc các file Excel...')
    readers = {}
    for key, src in excel_paths.items():
        readers[key] = ExcelReader(src)
        print(f'    OK: {key}')

    print('  [2/9] Tạo DocxModifier từ template...')
    doc = DocxModifier(TEMPLATE_PATH)

    print('  [3/9] Cập nhật MBB + FBB + MyTV...')
    update_mbb_fbb_mytv(doc, readers['mbb'], readers['fbb'], readers['mytv'])

    print('  [4/9] Cập nhật MLL...')
    week = update_mll(doc, readers['mll'])

    print('  [5/9] Cập nhật i-Speed...')
    update_ispeed(doc, readers['ispeed'])

    print('  [6/9] Cập nhật 5S...')
    update_5s(doc, readers['5s'])

    print('  [7/9] Cập nhật XLSC...')
    update_xlsc(doc, readers['xlsc'])

    print('  [8/9] Cập nhật Phụ lục...')
    update_appendix(doc, readers['appendix'])

    print('  [9/9] Cập nhật số tuần/ngày tháng...')
    replace_report_week(doc, week)

    print(f'  Lưu file kết quả: {output_path}')
    doc.save(output_path)
    print('  ✅ Hoàn thành!')
    return output_path


def generate_report_bytes(excel_paths_or_bytes: dict) -> bytes:
    """Tương tự generate_report nhưng trả về bytes thay vì lưu file."""
    readers = {k: ExcelReader(v) for k, v in excel_paths_or_bytes.items()}
    doc = DocxModifier(TEMPLATE_PATH)
    update_mbb_fbb_mytv(doc, readers['mbb'], readers['fbb'], readers['mytv'])
    week = update_mll(doc, readers['mll'])
    update_ispeed(doc, readers['ispeed'])
    update_5s(doc, readers['5s'])
    update_xlsc(doc, readers['xlsc'])
    update_appendix(doc, readers['appendix'])
    replace_report_week(doc, week)
    return doc.save_bytes()


if __name__ == '__main__':
    # Test mode: đọc từ thư mục local
    test_dir = os.path.join(os.path.dirname(__file__), '..', '..', '..', '..', 'Desktop')
    # Tùy chỉnh đường dẫn file test nếu cần
    excel_paths = {
        'mbb':     r'C:\Users\quyng\Desktop\1. BÁO CÁO MBB_HUNG.xlsx',
        'fbb':     r'C:\Users\quyng\Desktop\2. BÁO CÁO FBB_BAO.xlsx',
        'mytv':    r'C:\Users\quyng\Desktop\3. BÁO CÁO MYTV_TÂN.xlsx',
        'mll':     r'C:\Users\quyng\Desktop\4. BÁO CÁO MLL_KHANH.xlsx',
        'ispeed':  r'C:\Users\quyng\Desktop\5. BÁO CÁO ISPEED_QUOC.xlsx',
        '5s':      r'C:\Users\quyng\Desktop\6. BÁO CÁO 5S NHÀ TRẠM_TÂN.xlsx',
        'xlsc':    r'C:\Users\quyng\Desktop\7.BÁO CÁO XLSC_TUẤN.xlsx',
        'appendix': r'C:\Users\quyng\Desktop\PHỤ LỤC 1.xlsx',
    }
    missing = [k for k, v in excel_paths.items() if not os.path.exists(v)]
    if missing:
        print(f'Thiếu file: {missing}')
        print('Chạy script này từ API route thay vì trực tiếp.')
    else:
        output = r'C:\Users\quyng\Desktop\Bao_cao_VNPT_output.docx'
        generate_report(excel_paths, output)
