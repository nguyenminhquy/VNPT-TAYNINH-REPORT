import pandas as pd, openpyxl, sys, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from datetime import datetime
import re

script_dir = os.path.dirname(os.path.abspath(__file__))

def find_file(patterns):
    for fn in os.listdir(script_dir):
        fn_lower = fn.lower().replace(' ','_')
        if all(p.lower().replace(' ','_') in fn_lower for p in patterns):
            return os.path.join(script_dir, fn)
    return None

def find_files(patterns):
    """Giống find_file nhưng trả về TẤT CẢ file khớp mẫu (dùng cho Tiến trình
    xử lý sự cố vì hệ thống có thể xuất ra nhiều file con thay vì 1 file)."""
    matched = []
    for fn in os.listdir(script_dir):
        fn_lower = fn.lower().replace(' ', '_')
        if all(p.lower().replace(' ', '_') in fn_lower for p in patterns):
            matched.append(os.path.join(script_dir, fn))
    return sorted(matched)

xlsc_f      = find_file(['xlsc_brcd','chi_ti']) or find_file(['xlsc_brcd'])
export_f    = find_file(['export'])
tientrinh_fs= find_files(['n_tr']) or find_files(['n tr']) or find_files(['tien'])
votuyen_f   = find_file(['bao_cao']) or find_file(['bao cao'])
nv_f        = find_file(['nv_csht']) or find_file(['ds_nv'])
qlcsht_f    = find_file(['quản_lý_csht']) or find_file(['quan_ly_csht']) or find_file(['quản_lý_csht']) or find_file(['quanlycsht'])

missing=[]
if not xlsc_f: missing.append('XLSC_BRCD_CHI_TIET_CD5_...xlsx')
if not export_f: missing.append('..._export.xlsx')
if not tientrinh_fs: missing.append('Tien_trinh_xu_ly_su_co_...xlsx (1 hoac nhieu file)')
if not votuyen_f: missing.append('Bao_cao_XLSC_Tram_Vo_Tuyen_...xls')
if not nv_f: missing.append('DS_NV_CSHT...xlsx (mapping email nhan vien theo Ma CSHT)')
if not qlcsht_f: missing.append('Quản_lý_CSHT...xlsx (mapping Ten/SDT/To ha tang theo Ma CSHT)')
if missing:
    print('MISSING FILES:', missing); sys.exit(1)

print(f"XLSC: {os.path.basename(xlsc_f)}")
print(f"Export: {os.path.basename(export_f)}")
print(f"Tien trinh: {len(tientrinh_fs)} file -> " + ", ".join(os.path.basename(f) for f in tientrinh_fs))
print(f"Vo tuyen: {os.path.basename(votuyen_f)}")
print(f"NV CSHT (email): {os.path.basename(nv_f)}")
print(f"Quan ly CSHT (ten/sdt/to): {os.path.basename(qlcsht_f)}")

df1 = pd.read_excel(xlsc_f)
df2_raw = pd.read_excel(export_f, header=None)
df2 = df2_raw.iloc[2:].copy(); df2.columns = df2_raw.iloc[1].tolist(); df2 = df2.reset_index(drop=True)
to_ht = df2[df2['Tên Đơn vị Quản lý'].astype(str).str.contains('Tổ Hạ tầng', na=False)]
csht_to_to = dict(zip(to_ht['Mã CSHT'].astype(str), to_ht['Tên Đơn vị Quản lý'].astype(str)))

# Tra cứu Mã CSHT theo "Mã Khai thác" (cột D file export) — dùng để bù các dòng
# Vô Tuyến bị thiếu Mã CSHT (không tra được qua Mã phiếu -> Tiến trình xử lý sự cố).
# Chuẩn hóa theo ĐỘ DÀI mã (không phụ thuộc Hãng sản xuất, vì có trạm NOKIA nhưng
# lại đặt tên/mã kiểu Ericsson và ngược lại):
#   - Mã ngắn (<=7 ký tự, ví dụ "TCH008M"): giữ nguyên
#   - Mã dài (ví dụ "4G-KTU011M-LAN", "UL_BLU029M_LAN"): bỏ 3 ký tự đầu, lấy 7 ký tự tiếp theo
def _norm_ma_khai_thac(v):
    v = str(v).strip()
    if v in ('nan',''): return None
    return v if len(v) <= 7 else v[3:10]

df2['_MK'] = df2['Mã Khai thác'].apply(_norm_ma_khai_thac)
n_mk_dup = df2.dropna(subset=['_MK'])['_MK'].duplicated().sum()
if n_mk_dup > 0:
    print(f'[EXPORT] Canh bao: {n_mk_dup} gia tri "Ma Khai thac" (da chuan hoa) bi trung trong file export, da giu dong dau tien cho moi gia tri')
mk_dedup = df2.dropna(subset=['_MK']).drop_duplicates(subset='_MK', keep='first')
mk_to_csht = dict(zip(mk_dedup['_MK'], mk_dedup['Mã CSHT'].astype(str)))

df3 = pd.concat([pd.read_excel(f) for f in tientrinh_fs], ignore_index=True)
print(f"[TT] Tong {len(df3)} dong sau khi gop {len(tientrinh_fs)} file Tien trinh")
ma_sc_to_csht = dict(zip(df3['Mã sự cố'].astype(str), df3['Mã CSHT'].astype(str)))

# ── Mapping EMAIL nhân viên theo Mã CSHT (file DS_NV_CSHT) ───────────────
# Tự dò cột Mã CSHT (chứa "CSHT_") và cột Email (chứa "@") theo NỘI DUNG thay vì
# ép cứng số lượng/thứ tự cột, vì file nguồn có thể có số cột khác nhau.
nv_raw = pd.read_excel(nv_f, header=None)

def _find_col_by_content(df, predicate):
    best_col, best_ratio = None, 0
    for c in df.columns:
        vals = df[c].dropna().astype(str)
        if len(vals) == 0: continue
        ratio = predicate(vals).mean()
        if ratio > best_ratio:
            best_col, best_ratio = c, ratio
    return best_col if best_ratio > 0.3 else None

col_csht_nv  = _find_col_by_content(nv_raw, lambda s: s.str.contains('CSHT_', na=False))
col_email_nv = _find_col_by_content(nv_raw, lambda s: s.str.contains('@', na=False))
if col_csht_nv is None or col_email_nv is None:
    print(f'LOI: khong tim thay cot Ma CSHT hoac Email trong file {os.path.basename(nv_f)} '
          f'(tim thay Ma CSHT: {col_csht_nv is not None}, Email: {col_email_nv is not None})')
    sys.exit(1)
nv_raw = nv_raw.rename(columns={col_csht_nv: 'MA_CSHT_NV', col_email_nv: 'EMAIL_NV_RAW'})
nv_raw['MA_CSHT_NV'] = nv_raw['MA_CSHT_NV'].astype(str).str.strip()
nv_raw = nv_raw[nv_raw['MA_CSHT_NV'].str.contains('CSHT_', na=False)]

n_dup = nv_raw['MA_CSHT_NV'].duplicated().sum()
if n_dup > 0:
    print(f'[NV] Canh bao: {n_dup} Ma CSHT co nhieu nhan vien trong file DS_NV_CSHT, da gop het email (phan cach bang " / ")')

def _join_unique(s):
    seen = []
    for v in s:
        if pd.isna(v): continue
        v = str(v)
        if v not in seen: seen.append(v)
    return ' / '.join(seen)

nv_dedup = nv_raw.groupby('MA_CSHT_NV', as_index=False).agg({'EMAIL_NV_RAW': _join_unique})
csht_to_email = dict(zip(nv_dedup['MA_CSHT_NV'].astype(str), nv_dedup['EMAIL_NV_RAW']))

# ── Mapping TÊN / SĐT / TỔ HẠ TẦNG nhân viên theo Mã CSHT (file Quản lý CSHT) ──
ql_raw = pd.read_excel(qlcsht_f, header=1)
ql_raw['MA_CSHT_Q'] = ql_raw['Mã CSHT'].astype(str).str.strip()
n_dup_q = ql_raw['MA_CSHT_Q'].duplicated().sum()
if n_dup_q > 0:
    print(f'[QL_CSHT] Canh bao: {n_dup_q} Ma CSHT bi trung trong file Quan ly CSHT, da giu dong dau tien cho moi ma')
ql_dedup = ql_raw.drop_duplicates(subset='MA_CSHT_Q', keep='first')

def _norm_to_ht(v):
    v = '' if pd.isna(v) else str(v).strip()
    if v == '' or v.lower().startswith('tổ hạ tầng') or v.lower().startswith('to ha tang'):
        return v
    return f'Tổ Hạ tầng {v}'

csht_to_ten   = dict(zip(ql_dedup['MA_CSHT_Q'], ql_dedup['Kỹ thuật']))
csht_to_sdt   = dict(zip(ql_dedup['MA_CSHT_Q'], ql_dedup['SĐT KT']))
csht_to_to_nv = dict(zip(ql_dedup['MA_CSHT_Q'], ql_dedup['Tổ hạ tầng'].apply(_norm_to_ht)))

# ── Thời gian đóng phiếu (Kéo dài) ───────────────────────────────────────
def parse_keodai(s):
    if pd.isna(s): return None
    parts = str(s).strip().split(':')
    if len(parts) != 3: return None
    h, m, sec = parts
    if h == '' and m == '' and sec == '': return None
    try:
        return int(h or 0) + int(m or 0)/60 + int(sec or 0)/3600
    except ValueError:
        return None

df3['DUR_H']  = df3['Kéo dài'].apply(parse_keodai)
df3['EMAIL']  = df3['Mã CSHT'].astype(str).map(csht_to_email)
df3['TEN_NV'] = df3['Mã CSHT'].astype(str).map(csht_to_ten)
df3['SDT_NV'] = df3['Mã CSHT'].astype(str).map(csht_to_sdt)
df3['TO_NV']  = df3['Mã CSHT'].astype(str).map(csht_to_to_nv)

df4_raw = pd.read_excel(votuyen_f, header=None)
df4 = df4_raw.iloc[3:].copy(); df4.columns = df4_raw.iloc[2].tolist()
df4_before = len(df4)
mask_invalid = df4['Mã phiếu'].isna() | (df4['Mã phiếu'].astype(str).str.strip()=='') | (df4['Mã phiếu'].astype(str).str.strip()=='0')
dropped = df4[mask_invalid]
if len(dropped)>0:
    print(f'[VT] Bo {len(dropped)} dong Ma phieu rong/NaN/0 (index goc: {dropped.index.tolist()[:10]})')
df4 = df4[~mask_invalid].reset_index(drop=True)
print(f'[VT] Tong phieu sau loc: {len(df4)} (tu {df4_before} dong)')
df4['Mã CSHT'] = df4['Mã phiếu'].astype(str).map(ma_sc_to_csht)

# Bù các dòng thiếu Mã CSHT bằng cách tra "Tên node" -> Mã Khai thác (export),
# chuẩn hóa theo độ dài mã (xem hàm _norm_ma_khai_thac ở trên).
if 'Tên node' in df4.columns:
    def _derive_khaithac_key(ten_node):
        first = str(ten_node).split(',')[0].strip()
        return _norm_ma_khai_thac(first)

    mask_missing = df4['Mã CSHT'].isna() | (df4['Mã CSHT'].astype(str).str.strip().isin(['','nan']))
    n_missing_before = int(mask_missing.sum())
    key_node = df4['Tên node'].apply(_derive_khaithac_key)
    csht_tu_node = key_node.map(mk_to_csht)
    df4.loc[mask_missing, 'Mã CSHT'] = csht_tu_node[mask_missing]
    mask_still_missing = df4['Mã CSHT'].isna() | (df4['Mã CSHT'].astype(str).str.strip().isin(['','nan']))
    n_filled = n_missing_before - int(mask_still_missing.sum())
    print(f'[VT] Da bo sung {n_filled}/{n_missing_before} dong thieu Ma CSHT bang Ten node (chuan hoa do dai ma)')
else:
    print('[VT] Khong tim thay cot "Tên node", bo qua buoc bu Ma CSHT')

df4['TỔ HẠ TẦNG'] = df4['Mã CSHT'].astype(str).map(csht_to_to)
mask = df4['TỔ HẠ TẦNG'].isna() & df4['Đơn vị xử lý'].astype(str).str.contains('Tổ Hạ tầng', na=False)
df4.loc[mask,'TỔ HẠ TẦNG'] = df4.loc[mask,'Đơn vị xử lý']
df1['TỔ HẠ TẦNG'] = df1['MÃ CSHT'].astype(str).map(csht_to_to)

# Thông tin nhân viên (mapping theo Mã CSHT, từ file DS_NV_CSHT)
df1['TÊN NHÂN VIÊN']   = df1['MÃ CSHT'].astype(str).map(csht_to_ten)
df1['SĐT NHÂN VIÊN']   = df1['MÃ CSHT'].astype(str).map(csht_to_sdt)
df1['EMAIL NHÂN VIÊN'] = df1['MÃ CSHT'].astype(str).map(csht_to_email)
df1['TỔ HẠ TẦNG NV']   = df1['MÃ CSHT'].astype(str).map(csht_to_to_nv)

df4['TÊN NHÂN VIÊN']   = df4['Mã CSHT'].astype(str).map(csht_to_ten)
df4['SĐT NHÂN VIÊN']   = df4['Mã CSHT'].astype(str).map(csht_to_sdt)
df4['EMAIL NHÂN VIÊN'] = df4['Mã CSHT'].astype(str).map(csht_to_email)
df4['TỔ HẠ TẦNG NV']   = df4['Mã CSHT'].astype(str).map(csht_to_to_nv)

ALL_TO = sorted(['Tổ Hạ tầng Bến Lức','Tổ Hạ tầng Đức Hòa','Tổ Hạ tầng Gò Dầu','Tổ Hạ tầng Kiến Tường','Tổ Hạ tầng Tân An','Tổ Hạ tầng Tân Châu','Tổ Hạ tầng Tân Ninh'])

def pct(a,b): return round(a/b*100,2) if b>0 else 0.0  # float 2 số lẻ

def stats_new(df, to_col, tt_col, qh_col):
    rows=[]
    for to in ALL_TO:
        sub=df[df[to_col]==to]; total=len(sub)
        ht=int((sub[tt_col].astype(str)=='Đã xác nhận').sum())
        ht_dh=int(((sub[tt_col].astype(str)=='Đã xác nhận')&(sub[qh_col].astype(str).str.strip()=='No')).sum())
        ht_qh=int(((sub[tt_col].astype(str)=='Đã xác nhận')&(sub[qh_col].astype(str).str.strip()=='Yes')).sum())
        ton=total-ht
        ton_th=int(((sub[tt_col].astype(str)!='Đã xác nhận')&(sub[qh_col].astype(str).str.strip()=='No')).sum())
        ton_qh=int(((sub[tt_col].astype(str)!='Đã xác nhận')&(sub[qh_col].astype(str).str.strip()=='Yes')).sum())
        rows.append({'Tổ Hạ tầng':to,'Tổng giao':int(total),'Hoàn thành':ht,'HT đúng hạn':ht_dh,'HT quá hạn':ht_qh,'Tỉ lệ HT đúng hạn':pct(ht_dh,ht),'Tổng tồn':int(ton),'Tồn trong hạn':ton_th,'Tồn quá hạn':ton_qh,'Tỉ lệ tồn QH':pct(int(ton),int(total))})
    df_o=pd.DataFrame(rows)
    t_giao=int(df_o['Tổng giao'].sum()); t_ht=int(df_o['Hoàn thành'].sum()); t_htdh=int(df_o['HT đúng hạn'].sum()); t_ton=int(df_o['Tổng tồn'].sum()); t_tonqh=int(df_o['Tồn quá hạn'].sum())
    tr={'Tổ Hạ tầng':'TỔNG','Tổng giao':t_giao,'Hoàn thành':t_ht,'HT đúng hạn':t_htdh,'HT quá hạn':int(df_o['HT quá hạn'].sum()),'Tỉ lệ HT đúng hạn':pct(t_htdh,t_ht),'Tổng tồn':t_ton,'Tồn trong hạn':int(df_o['Tồn trong hạn'].sum()),'Tồn quá hạn':t_tonqh,'Tỉ lệ tồn QH':pct(t_ton,t_giao)}
    return pd.concat([df_o,pd.DataFrame([tr])],ignore_index=True)

mane=df1[df1['LOẠI MẠNG']=='MANE']; access_df=df1[df1['LOẠI MẠNG']=='ACCESS']
stat_mane=stats_new(mane,'TỔ HẠ TẦNG','TRẠNG THÁI','PHIẾU QUÁ HẠN TOÀN TRÌNH')
stat_access=stats_new(access_df,'TỔ HẠ TẦNG','TRẠNG THÁI','PHIẾU QUÁ HẠN TOÀN TRÌNH')
stat_vt=stats_new(df4,'TỔ HẠ TẦNG','Trạng thái','Phiếu quá hạn toàn trình')

dates=re.findall(r'(\d{2}-\d{2}-\d{4})',os.path.basename(xlsc_f))
date_label=f"{dates[0]} - {dates[1]}" if len(dates)>=2 else datetime.today().strftime('%d/%m/%Y')

# ── MÀU SẮC ────────────────────────────────────────────────────
CH_DARK="1F3864"; CH_MID="2E75B6"; CM="4472C4"; CA="ED7D31"; CV="70AD47"
CQH="C00000"; CTH="375623"; CTOT="D6E4F0"; CODD="EBF3FB"; CEVN="FFFFFF"
# Màu phiếu mới
C_TON_QH="BF8F00"    # Vàng đậm - tồn quá hạn - chữ trắng
C_HT_QH ="A50000"    # Đỏ đậm - HT quá hạn - chữ trắng (giữ nguyên)
C_TON_TH="1F4E79"    # Blue đậm - tồn trong hạn - chữ trắng

def tb(): s=Side(style='thin',color='AAAAAA'); return Border(left=s,right=s,top=s,bottom=s)
def hd(cell,bg,fc="FFFFFF",sz=10,wrap=False,left=False):
    cell.fill=PatternFill("solid",fgColor=bg); cell.font=Font(bold=True,color=fc,name='Arial',size=sz)
    cell.alignment=Alignment(horizontal='left' if left else 'center',vertical='center',wrap_text=wrap); cell.border=tb()
def dt(cell,bg=None,bold=False,fc="000000",left=False):
    if bg: cell.fill=PatternFill("solid",fgColor=bg)
    cell.font=Font(bold=bold,color=fc,name='Arial',size=10)
    cell.alignment=Alignment(horizontal='left' if left else 'center',vertical='center'); cell.border=tb()

wb=openpyxl.Workbook()
print("OPENPYXL =", openpyxl.__version__)
# Xóa sheet trắng mặc định
wb.remove(wb.active)

# ── BẢNG THỐNG KÊ 3 LOẠI (MANE / ACCESS / VÔ TUYẾN) ─────────
STAT_COLS=['Tổng giao','Hoàn thành','HT đúng hạn','HT quá hạn','Tỉ lệ HT đúng hạn','Tổng tồn','Tồn trong hạn','Tồn quá hạn','Tỉ lệ tồn QH']
STAT_HDRS=['Tổng\ngiao','Hoàn\nthành','HT đúng\nhạn','HT quá\nhạn','TL HT\nđúng hạn','Tổng\ntồn','Tồn\ntrong hạn','Tồn\nquá hạn','TL tồn\nQH']
STAT_COLORS=[CH_MID,CTH,CTH,CQH,CTH,CH_MID,"375623",CQH,CQH]

def write_stat_block(ws, start_row, title, stat_df, color_accent):
    ws.row_dimensions[start_row].height=28
    ws.merge_cells(start_row=start_row,start_column=1,end_row=start_row,end_column=len(STAT_COLS)+1)
    hd(ws.cell(row=start_row,column=1,value=title),color_accent,sz=11,left=True)
    hr=start_row+1; ws.row_dimensions[hr].height=40
    hd(ws.cell(row=hr,column=1,value='Tổ Hạ tầng'),CH_DARK,sz=10,wrap=True)
    for ci,(h,c) in enumerate(zip(STAT_HDRS,STAT_COLORS),start=2): hd(ws.cell(row=hr,column=ci,value=h),c,sz=9,wrap=True)
    for ri,row in stat_df.iterrows():
        r=hr+1+ri; ws.row_dimensions[r].height=22
        is_tot=row['Tổ Hạ tầng']=='TỔNG'; bg=CTOT if is_tot else (CODD if ri%2==0 else CEVN)
        dt(ws.cell(row=r,column=1,value=row['Tổ Hạ tầng']),bg,bold=is_tot,left=True)
        for ci,col in enumerate(STAT_COLS,start=2):
            val=row[col]; cell=ws.cell(row=r,column=ci,value=val)
            is_qhcol=(col in ['HT quá hạn','Tồn quá hạn','Tỉ lệ tồn QH']); is_goodcol=(col in ['HT đúng hạn','Tỉ lệ HT đúng hạn','Tồn trong hạn'])
            fc="000000"
            if is_qhcol and not is_tot and isinstance(val,(int,float)) and val>0: fc=CQH
            elif is_goodcol and not is_tot: fc=CTH
            dt(cell,bg,bold=(is_tot or (is_qhcol and fc==CQH)),fc=fc)
            if col in ['Tỉ lệ HT đúng hạn','Tỉ lệ tồn QH']: cell.number_format='0.00"%"' 
    return hr+1+len(stat_df)

# ── SHEET XLSC MANE / ACCESS / VÔ TUYẾN ──────────────────────
def autofit_col(ws, col_idx, min_w=6, max_w=50, skip_wrapped=True, min_row=1, max_row=None):
    """Tính độ rộng tối ưu trong vùng hàng chỉ định (min_row..max_row)."""
    best=min_w
    col_letter=get_column_letter(col_idx)
    end_row = max_row if max_row else ws.max_row
    for r in range(min_row, end_row+1):
        cell=ws.cell(row=r, column=col_idx)
        if cell.value is None: continue
        if skip_wrapped and cell.alignment and cell.alignment.wrap_text: continue
        txt=str(cell.value)
        longest=max((len(ln) for ln in txt.split('\n')),default=0) if '\n' in txt else len(txt)
        w=longest*1.1+2
        if w>best: best=w
    ws.column_dimensions[col_letter].width=min(max(best,min_w),max_w)

def apply_color_scale(ws, col_letter, min_row, max_row, toward_zero=False):
    """Color scale: xanh lá tiến về tốt, đỏ tiến về xấu.
    Nếu toàn bộ giá trị = 0 (toward_zero) thì tô xanh hết, bỏ color scale."""
    all_zero = all(
        (ws.cell(row=r, column=ws[col_letter+str(r)].column).value or 0) == 0
        for r in range(min_row, max_row+1)
    )
    if toward_zero and all_zero:
        from openpyxl.styles import PatternFill
        for r in range(min_row, max_row+1):
            c = ws[f'{col_letter}{r}']
            c.fill = PatternFill("solid", fgColor="63BE7B")
        return
    if toward_zero:
        rule=ColorScaleRule(
            start_type='num',  start_value=0,   start_color='63BE7B',
            mid_type='num',    mid_value=5,      mid_color='FFEB84',
            end_type='max',    end_value=None,   end_color='F8696B')
    else:
        rule=ColorScaleRule(
            start_type='num',  start_value=0,   start_color='F8696B',
            mid_type='num',    mid_value=80,     mid_color='FFEB84',
            end_type='num',    end_value=100,    end_color='63BE7B')
    ws.conditional_formatting.add(f'{col_letter}{min_row}:{col_letter}{max_row}',rule)

def write_sheet(ws,title,stat_df,detail_df,color_accent,qh_col,tt_col,src_cols,hdr_cols,widths,left_idx,stat_col_widths=None):
    ws.sheet_view.showGridLines=False
    n_stat_cols=len(STAT_COLS)+1
    ws.merge_cells(f'A1:{get_column_letter(n_stat_cols)}1'); ws['A1']=title; hd(ws['A1'],CH_DARK,sz=13); ws.row_dimensions[1].height=32
    ws.row_dimensions[2].height=42; hd(ws.cell(row=2,column=1,value='Tổ Hạ tầng'),CH_DARK,sz=10,wrap=True)
    for ci,(h,c) in enumerate(zip(STAT_HDRS,STAT_COLORS),start=2): hd(ws.cell(row=2,column=ci,value=h),c,sz=9,wrap=True)
    n_data_rows=len(stat_df)
    for ri,row in stat_df.iterrows():
        r=ri+3; ws.row_dimensions[r].height=22; is_tot=row['Tổ Hạ tầng']=='TỔNG'; bg=CTOT if is_tot else (CODD if ri%2==0 else CEVN)
        dt(ws.cell(row=r,column=1,value=row['Tổ Hạ tầng']),bg,bold=is_tot,left=True)
        for ci,col in enumerate(STAT_COLS,start=2):
            val=row[col]; cell=ws.cell(row=r,column=ci,value=val)
            is_qhcol=(col in ['HT quá hạn','Tồn quá hạn','Tỉ lệ tồn QH']); is_goodcol=(col in ['HT đúng hạn','Tỉ lệ HT đúng hạn','Tồn trong hạn'])
            fc="000000"
            if is_qhcol and not is_tot and isinstance(cell.value,(int,float)) and cell.value>0: fc=CQH
            elif is_goodcol and not is_tot: fc=CTH
            dt(cell,bg,bold=(is_tot or (is_qhcol and fc==CQH)),fc=fc)
            if col in ['Tỉ lệ HT đúng hạn','Tỉ lệ tồn QH']: cell.number_format='0.00"%"'

    data_start=3; data_end=3+n_data_rows-2
    col_tl_ht=get_column_letter(6)
    col_tl_ton=get_column_letter(10)
    if data_end>=data_start:
        apply_color_scale(ws,col_tl_ht, data_start,data_end,toward_zero=False)
        apply_color_scale(ws,col_tl_ton,data_start,data_end,toward_zero=True)

    sr=len(stat_df)+5

    # ── BẢNG CHÚ THÍCH ──────────────────────────────────────────
    ws.row_dimensions[sr].height=20
    ws.merge_cells(f'A{sr}:{get_column_letter(len(hdr_cols))}{sr}')
    ws.cell(row=sr,column=1,value='CHÚ THÍCH MÀU SẮC PHIẾU')
    hd(ws.cell(row=sr,column=1),'555555',sz=10,left=True)
    sr+=1
    legend=[
        (C_TON_QH,"FFFFFF","Tồn quá hạn (chưa xử lý + quá hạn)"),
        (C_HT_QH, "FFFFFF","Hoàn thành quá hạn (đã xử lý nhưng trễ hạn)"),
        (C_TON_TH,"FFFFFF","Tồn trong hạn (chưa xử lý + còn hạn)"),
        (CODD,    "000000","Hoàn thành đúng hạn"),
    ]
    for bg_l,fc_l,label_l in legend:
        ws.row_dimensions[sr].height=18
        c1=ws.cell(row=sr,column=1,value=f'   {label_l}')
        c1.fill=PatternFill("solid",fgColor=bg_l); c1.font=Font(color=fc_l,name='Arial',size=10,bold=True)
        c1.alignment=Alignment(horizontal='left',vertical='center'); c1.border=tb()
        ws.merge_cells(start_row=sr,start_column=1,end_row=sr,end_column=4)
        for ci2 in range(2,5):
            c2=ws.cell(row=sr,column=ci2)
            c2.fill=PatternFill("solid",fgColor=bg_l); c2.border=tb()
        sr+=1
    sr+=1

    # ── CHI TIẾT PHIẾU ──────────────────────────────────────────
    ws.merge_cells(f'A{sr}:{get_column_letter(len(hdr_cols))}{sr}')
    ws.cell(row=sr,column=1,value='▶  CHI TIẾT PHIẾU'); hd(ws.cell(row=sr,column=1),color_accent,sz=11,left=True); ws.row_dimensions[sr].height=26
    hr=sr+1; ws.row_dimensions[hr].height=34
    for ci,h in enumerate(hdr_cols,start=1): hd(ws.cell(row=hr,column=ci,value=h),color_accent,sz=9,wrap=True)
    for ri2,row2 in detail_df.reset_index(drop=True).iterrows():
        r2=hr+1+ri2; ws.row_dimensions[r2].height=20
        is_ton=(str(row2.get(tt_col,''))!='Đã xác nhận'); is_qh=(str(row2.get(qh_col,''))=='Yes')
        is_ht_qh=(not is_ton and is_qh)
        if is_ton and is_qh:   bg2,fc2=C_TON_QH,"FFFFFF"
        elif is_ht_qh:         bg2,fc2=C_HT_QH, "FFFFFF"
        elif is_ton:           bg2,fc2=C_TON_TH,"FFFFFF"
        else:                  bg2,fc2=(CODD if ri2%2==0 else CEVN),"000000"
        for ci,col in enumerate(src_cols,start=1):
            val=row2.get(col,''); cell=ws.cell(row=r2,column=ci,value=str(val) if pd.notna(val) else '')
            dt(cell,bg2,fc=fc2,left=(ci in left_idx))
    detail_end_row = hr + len(detail_df.reset_index(drop=True))
    for ci in range(1,len(hdr_cols)+1):
        autofit_col(ws,ci,min_w=8,max_w=40, min_row=hr,max_row=detail_end_row)
    _STAT_W = stat_col_widths if stat_col_widths else [21, 5.5, 6, 8, 7, 8.5, 5.5, 6, 7.5, 6.5]
    for _ci, _w in enumerate(_STAT_W, start=1):
        ws.column_dimensions[get_column_letter(_ci)].width = _w

sc=['MÃ PHIẾU','TRẠNG THÁI','LOẠI CẢNH BÁO','THỜI GIAN BẮT ĐẦU','THỜI GIAN KẾT THÚC','PHIẾU QUÁ HẠN TOÀN TRÌNH','ĐƠN VỊ XỬ LÝ','LOẠI MẠNG','LOẠI ĐỐI TƯỢNG','TÊN ĐỐI TƯỢNG','MÃ CSHT','TÊN NHÀ TRẠM','QUY TRÌNH','TỔ HẠ TẦNG','TÊN NHÂN VIÊN','SĐT NHÂN VIÊN','EMAIL NHÂN VIÊN','TỔ HẠ TẦNG NV']
hc=['Mã phiếu','Trạng thái','Loại cảnh báo','Bắt đầu','Kết thúc','Quá hạn TT','Đơn vị xử lý','Loại mạng','Loại đối tượng','Tên đối tượng','Mã CSHT','Tên nhà trạm','Quy trình','Tổ Hạ tầng','Tên nhân viên','SĐT nhân viên','Email nhân viên','Tổ HT (NV)']
mw=[12,18,28,20,20,10,20,10,14,26,16,20,14,24,22,16,28,22]
ws_m=wb.create_sheet("XLSC MANE"); ws_m.sheet_properties.tabColor=CM
write_sheet(ws_m,f'XLSC MANE - TÂY NINH ({date_label})',stat_mane,mane,CM,'PHIẾU QUÁ HẠN TOÀN TRÌNH','TRẠNG THÁI',sc,hc,mw,[3,7,10,11,12,15,16,17,18])
ws_a=wb.create_sheet("XLSC ACCESS"); ws_a.sheet_properties.tabColor=CA
write_sheet(ws_a,f'XLSC ACCESS - TÂY NINH ({date_label})',stat_access,access_df,CA,'PHIẾU QUÁ HẠN TOÀN TRÌNH','TRẠNG THÁI',sc,hc,mw,[3,7,10,11,12,15,16,17,18])
vsc=['Mã phiếu','Trạng thái','Loại sự cố','Thời gian bắt đầu','Thời gian kết thúc','Phiếu quá hạn toàn trình','Đơn vị xử lý','Loại node','Tên node','Loại trạm','Mã CSHT','TỔ HẠ TẦNG','TÊN NHÂN VIÊN','SĐT NHÂN VIÊN','EMAIL NHÂN VIÊN','TỔ HẠ TẦNG NV']
vhc=['Mã phiếu','Trạng thái','Loại sự cố','Bắt đầu','Kết thúc','Quá hạn TT','Đơn vị xử lý','Loại node','Tên node','Loại trạm','Mã CSHT','Tổ Hạ tầng','Tên nhân viên','SĐT nhân viên','Email nhân viên','Tổ HT (NV)']
vw=[12,18,14,20,20,12,22,12,20,10,16,24,22,16,28,22]
ws_v=wb.create_sheet("XLSC VÔ TUYẾN"); ws_v.sheet_properties.tabColor=CV
write_sheet(ws_v,f'XLSC VÔ TUYẾN - TÂY NINH ({date_label})',stat_vt,df4,CV,'Phiếu quá hạn toàn trình','Trạng thái',vsc,vhc,vw,[3,7,9,12,13,14,15,16],stat_col_widths=[21,5.5,6,8,7,8.5,5.5,9,7.5,6.5])

# ── SHEET PHIẾU QUÁ HẠN ──────────────────────────────────────
ws_qh = wb.create_sheet("PHIẾU QUÁ HẠN"); ws_qh.sheet_properties.tabColor="C00000"

mane_qh   = mane[mane['PHIẾU QUÁ HẠN TOÀN TRÌNH'].astype(str).str.strip()=='Yes'].copy()
access_qh = access_df[access_df['PHIẾU QUÁ HẠN TOÀN TRÌNH'].astype(str).str.strip()=='Yes'].copy()
vt_qh     = df4[df4['Phiếu quá hạn toàn trình'].astype(str).str.strip()=='Yes'].copy()

qh_src_ma = ['MÃ PHIẾU','TRẠNG THÁI','LOẠI CẢNH BÁO','THỜI GIAN BẮT ĐẦU','THỜI GIAN KẾT THÚC','ĐƠN VỊ XỬ LÝ','LOẠI MẠNG','TÊN ĐỐI TƯỢNG','TỔ HẠ TẦNG','EMAIL NHÂN VIÊN']
qh_hc_ma  = ['Mã phiếu','Trạng thái','Loại cảnh báo','Bắt đầu','Kết thúc','Đơn vị xử lý','Loại mạng','Tên đối tượng','Tổ Hạ tầng','Email nhân viên']
qh_src_vt = ['Mã phiếu','Trạng thái','Loại sự cố','Thời gian bắt đầu','Thời gian kết thúc','Đơn vị xử lý','Loại node','Tên node','TỔ HẠ TẦNG','EMAIL NHÂN VIÊN']
qh_hc_vt  = ['Mã phiếu','Trạng thái','Loại sự cố','Bắt đầu','Kết thúc','Đơn vị xử lý','Loại node','Tên node','Tổ Hạ tầng','Email nhân viên']
qh_w      = [12,18,22,20,20,22,14,26,24,28]

def write_qh_section(ws, start_row, label, color_tab, df_sub, src_cols, hdr_cols, left_idx):
    ws.row_dimensions[start_row].height = 24
    cell = ws.cell(row=start_row, column=1, value=f"  {label}  \u2014  {len(df_sub)} phi\u1ebfu qu\u00e1 h\u1ea1n")
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill("solid", fgColor=color_tab)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(hdr_cols))
    hr = start_row + 1
    ws.row_dimensions[hr].height = 20
    for ci, h in enumerate(hdr_cols, start=1):
        ch2 = ws.cell(row=hr, column=ci, value=h)
        dt(ch2, "7B0000", bold=True, fc="FFFFFF", left=(ci in left_idx))
    if len(df_sub) == 0:
        ws.cell(row=hr+1, column=1, value="(Kh\u00f4ng c\u00f3 phi\u1ebfu qu\u00e1 h\u1ea1n)")
        return hr + 3
    for ri, (_, row) in enumerate(df_sub.iterrows()):
        dr = hr + 1 + ri
        ws.row_dimensions[dr].height = 18
        for ci, col in enumerate(src_cols, start=1):
            val = row.get(col, '')
            cell = ws.cell(row=dr, column=ci, value=str(val) if pd.notna(val) else '')
            dt(cell, CEVN, fc="000000", left=(ci in left_idx))
    return hr + 1 + len(df_sub) + 2

ws_qh.row_dimensions[1].height = 30
t = ws_qh.cell(row=1, column=1, value="DANH SÁCH PHIẾU QUÁ HẠN - TÂY NINH (" + date_label + ")")
t.font = Font(bold=True, color="FFFFFF", size=14)
t.fill = PatternFill("solid", fgColor="7B0000")
t.alignment = Alignment(horizontal="center", vertical="center")
ws_qh.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(qh_hc_ma))

ws_qh.row_dimensions[2].height = 18
summary = "Tổng quá hạn: MANE " + str(len(mane_qh)) + " | ACCESS " + str(len(access_qh)) + " | Vô tuyến " + str(len(vt_qh)) + " | Tổng cộng " + str(len(mane_qh)+len(access_qh)+len(vt_qh)) + " phiếu"
sc2 = ws_qh.cell(row=2, column=1, value=summary)
sc2.font = Font(bold=True, color="7B0000", size=11)
sc2.fill = PatternFill("solid", fgColor="FFE0E0")
sc2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_qh.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(qh_hc_ma))

cur = 4
cur = write_qh_section(ws_qh, cur, "MANE",       CM,  mane_qh,   qh_src_ma, qh_hc_ma, [3,7,8,10])
cur = write_qh_section(ws_qh, cur, "ACCESS",     CA,  access_qh, qh_src_ma, qh_hc_ma, [3,7,8,10])
cur = write_qh_section(ws_qh, cur, "VÔ TUYẾN",  CV,  vt_qh,     qh_src_vt, qh_hc_vt, [3,7,8,10])

for ci, w in enumerate(qh_w, start=1):
    ws_qh.column_dimensions[get_column_letter(ci)].width = w
ws_qh.column_dimensions['A'].width=21   # Tổ HT
ws_qh.freeze_panes = "A3"

# ── SHEET THỐNG KÊ EMAIL NHÂN VIÊN (chậm nhất → nhanh nhất) ──
closed = df3[df3['Trạng thái sự cố'] == 'Đã nghiệm thu']
valid = closed.dropna(subset=['EMAIL', 'DUR_H'])
agg = valid.groupby('EMAIL').agg(
    so_phieu=('DUR_H', 'count'),
    TB_RAW=('DUR_H', 'mean'),
    Ten=('TEN_NV', 'first'),
    SDT=('SDT_NV', 'first'),
    To=('TO_NV', 'first'),
).reset_index()
agg = agg.rename(columns={'EMAIL':'Email','so_phieu':'Số phiếu'})
agg = agg.sort_values('TB_RAW', ascending=False).reset_index(drop=True)

# Chốt về SỐ PHÚT để TB giờ làm tròn nhất quán (không phụ thuộc lần làm tròn khác).
agg['TOTAL_MIN'] = agg['TB_RAW'].apply(lambda h: round(h * 60))
agg['TB giờ'] = (agg['TOTAL_MIN'] / 60).round(2)

ws_s = wb.create_sheet('TB ĐÓNG PHIẾU TỔNG HỢP'); ws_s.sheet_properties.tabColor = "7030A0"
ws_s.sheet_view.showGridLines = False
ncols = 7
ws_s.merge_cells(f'A1:{get_column_letter(ncols)}1')
ws_s['A1'] = f'TB THỜI GIAN ĐÓNG PHIẾU TỔNG HỢP (MANE + ACCESS + VÔ TUYẾN) ({date_label})'
hd(ws_s['A1'], CH_DARK, sz=12, wrap=True); ws_s.row_dimensions[1].height = 30

hr = 3
ws_s.row_dimensions[hr].height = 26
headers = ['STT', 'Email nhân viên', 'Tên nhân viên', 'SĐT nhân viên', 'Tổ hạ tầng', 'Số phiếu xử lý', 'TB thời gian đóng phiếu (giờ)']
for ci, h in enumerate(headers, start=1):
    hd(ws_s.cell(row=hr, column=ci, value=h), CH_DARK, sz=10, wrap=True)

for i, row in agg.iterrows():
    r = hr + 1 + i
    bg = CODD if i % 2 == 0 else CEVN
    ws_s.row_dimensions[r].height = 20
    dt(ws_s.cell(row=r, column=1, value=i + 1), bg)
    dt(ws_s.cell(row=r, column=2, value=row['Email']), bg, left=True)
    dt(ws_s.cell(row=r, column=3, value=row['Ten']), bg, left=True)
    dt(ws_s.cell(row=r, column=4, value=row['SDT']), bg, left=True)
    dt(ws_s.cell(row=r, column=5, value=row['To']), bg, left=True)
    dt(ws_s.cell(row=r, column=6, value=int(row['Số phiếu'])), bg)
    dt(ws_s.cell(row=r, column=7, value=float(row['TB giờ'])), bg)

data_start = hr + 1; data_end = hr + len(agg)
if data_end >= data_start:
    rule = ColorScaleRule(
        start_type='min', start_color='63BE7B',
        mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max', end_color='F8696B')
    ws_s.conditional_formatting.add(f'G{data_start}:G{data_end}', rule)

widths = [6, 32, 26, 16, 22, 14, 26]
for ci, w in enumerate(widths, start=1):
    ws_s.column_dimensions[get_column_letter(ci)].width = w
ws_s.freeze_panes = f'A{hr+1}'
print(f'[EMAIL] Sheet TB DONG PHIEU TONG HOP: {len(agg)} nhan vien')

# ── SHEET TB THỜI GIAN ĐÓNG PHIẾU THEO LOẠI: MANE / ACCESS / VÔ TUYẾN ──────
# Chuẩn hóa key: loại bỏ .0 (float) khi convert sang str
def norm_ma(s):
    s = str(s).strip()
    return s[:-2] if s.endswith('.0') else s

ma_sc_to_loai = {}
for _, r in df1.iterrows():
    ma_sc_to_loai[norm_ma(r['MÃ PHIẾU'])] = str(r.get('LOẠI MẠNG', ''))
for _, r in df4.iterrows():
    ma_sc_to_loai[norm_ma(r['Mã phiếu'])] = 'VÔ TUYẾN'

df3['LOAI_MANG'] = df3['Mã sự cố'].apply(norm_ma).map(ma_sc_to_loai)

def make_email_sheet(ws_out, title_full, loai_filter, tab_color, date_lbl):
    """Tạo sheet TB thời gian đóng phiếu lọc theo loại mạng."""
    subset   = df3[df3['LOAI_MANG'] == loai_filter]
    closed_s = subset[subset['Trạng thái sự cố'] == 'Đã nghiệm thu']
    valid_s  = closed_s.dropna(subset=['EMAIL', 'DUR_H'])
    agg_s    = valid_s.groupby('EMAIL').agg(
        so_phieu=('DUR_H', 'count'),
        TB_RAW=('DUR_H', 'mean'),
        Ten=('TEN_NV', 'first'),
        SDT=('SDT_NV', 'first'),
        To=('TO_NV', 'first'),
    ).reset_index()
    agg_s    = agg_s.rename(columns={'EMAIL':'Email','so_phieu':'Số phiếu'})
    agg_s    = agg_s.sort_values('TB_RAW', ascending=False).reset_index(drop=True)
    agg_s['TOTAL_MIN'] = agg_s['TB_RAW'].apply(lambda h: round(h * 60))
    agg_s['TB giờ']    = (agg_s['TOTAL_MIN'] / 60).round(2)

    ws_out.sheet_properties.tabColor = tab_color
    ws_out.sheet_view.showGridLines   = False
    ncols_s = 7
    ws_out.merge_cells(f'A1:{get_column_letter(ncols_s)}1')
    ws_out['A1'] = f'{title_full} ({date_lbl})'
    hd(ws_out['A1'], CH_DARK, sz=12, wrap=True)
    ws_out.row_dimensions[1].height = 30

    hr_s = 3
    ws_out.row_dimensions[hr_s].height = 26
    headers_s = ['STT', 'Email nhân viên', 'Tên nhân viên', 'SĐT nhân viên', 'Tổ hạ tầng', 'Số phiếu xử lý', 'TB thời gian đóng phiếu (giờ)']
    for ci, h in enumerate(headers_s, start=1):
        hd(ws_out.cell(row=hr_s, column=ci, value=h), CH_DARK, sz=10, wrap=True)

    for i, row in agg_s.iterrows():
        r = hr_s + 1 + i
        bg = CODD if i % 2 == 0 else CEVN
        ws_out.row_dimensions[r].height = 20
        dt(ws_out.cell(row=r, column=1, value=i + 1), bg)
        dt(ws_out.cell(row=r, column=2, value=row['Email']), bg, left=True)
        dt(ws_out.cell(row=r, column=3, value=row['Ten']), bg, left=True)
        dt(ws_out.cell(row=r, column=4, value=row['SDT']), bg, left=True)
        dt(ws_out.cell(row=r, column=5, value=row['To']), bg, left=True)
        dt(ws_out.cell(row=r, column=6, value=int(row['Số phiếu'])), bg)
        dt(ws_out.cell(row=r, column=7, value=float(row['TB giờ'])), bg)

    ds = hr_s + 1; de = hr_s + len(agg_s)
    if de >= ds:
        rule_s = ColorScaleRule(
            start_type='min',        start_color='63BE7B',
            mid_type='percentile',   mid_value=50, mid_color='FFEB84',
            end_type='max',          end_color='F8696B')
        ws_out.conditional_formatting.add(f'G{ds}:G{de}', rule_s)

    for ci, w in enumerate([6, 32, 26, 16, 22, 14, 26], start=1):
        ws_out.column_dimensions[get_column_letter(ci)].width = w
    ws_out.freeze_panes = f'A{hr_s+1}'
    print(f'[EMAIL] Sheet {ws_out.title}: {len(agg_s)} nhan vien, {len(valid_s)} phieu')

ws_sm = wb.create_sheet('TB ĐÓNG PHIẾU MANE')
make_email_sheet(ws_sm, 'TB THỜI GIAN ĐÓNG PHIẾU - MANE', 'MANE', CM, date_label)
ws_sa = wb.create_sheet('TB ĐÓNG PHIẾU ACCESS')
make_email_sheet(ws_sa, 'TB THỜI GIAN ĐÓNG PHIẾU - ACCESS', 'ACCESS', CA, date_label)
ws_sv = wb.create_sheet('TB ĐÓNG PHIẾU VÔ TUYẾN')
make_email_sheet(ws_sv, 'TB THỜI GIAN ĐÓNG PHIẾU - VÔ TUYẾN', 'VÔ TUYẾN', CV, date_label)

out=os.path.join(script_dir,'BaoCao_XLSC_TayNinh_Updated.xlsx')
if os.path.exists(out):
    os.remove(out)
wb.save(out)
print(f"SUCCESS - Saved: {out}")
